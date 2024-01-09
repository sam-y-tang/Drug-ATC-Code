from openpyxl import load_workbook
import csv


def process(excelFile, csvFile):
    wb = load_workbook(excelFile)

    sheet_obj = wb.active
    max_col = sheet_obj.max_column

    # Loop will print all columns name
    # for i in range(1, max_col + 1):
    #     cell_obj = sheet_obj.cell(row=1, column=i)
    #     print(cell_obj.value)


    code_dict = dict()
    with open(csvFile, mode='r') as file:
        csv_reader = csv.DictReader(file)

        # converting the file to dictionary
        # by first converting to list
        # and then converting the list to dict
        print(csv_reader.fieldnames)

        for row in list(csv_reader):
            code=row['ATC.code']
            if len(code)==7:
                code_dict[row['ATC.code']]=row['Name'].upper()

    cnt=0

    # for k, v in code_dict.items():
    #     if cnt<5:
    #         print(k, v)
    #     cnt+=1

    cnt=0
    sheet_obj.insert_cols(3)
    for i in range(1, sheet_obj.max_row + 1):

        print("Row %4d "% i, end="")
        name = sheet_obj.cell(row=i, column=2).value
        code = find_code(name, code_dict)
        print(name, code)
        if len(code)==0:
            cnt+=1
        else:
            sheet_obj.cell(row=i, column=3).value=code

    print("Total %d lines have no corresponding ATC code"%cnt)
    import os.path
    base =os.path.basename(excelFile)
    filename,ext = os.path.splitext(base)
    filename+="_new"+ext
    sheet_obj.cell(row=1, column=3).value = "ATC.code"
    wb.save(filename)
    print("add code column and save to file "+filename)
def find_code(name, code_dict):
    code=[]
    name=name.upper()
    name=name.split(',')
    for gredient in name:
        for k,v in code_dict.items():
            #print(k, v)
            if v.find(gredient)>=0:
                code.append(k)

    return ",".join(code)


if __name__ == "__main__":
    process("List of drugs 231214.xlsx", "atccodes.csv")
